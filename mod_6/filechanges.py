import os
import sys
import hashlib
import socket
import sqlite3
from datetime import datetime
from openpyxl import workbook
from openpyxl.styles import Font

def basefile():
    """ Return name of base file. """
    return os.path.splitext(os.path.basename(__file__))[0]

def connectDb():
    """ Connect to core database. Database will be created if it doens't already exist. """
    database = basefile() + ".db"

    try:
        conn = sqlite3.connect(database, timeout=2)
    except BaseException as err:
        print(str(err))
        conn = None

    return conn

def coreCursor(conn, query, args):
    """ Run query against connected database. """
    result = False
    cursor = conn.cursor()
    try:
        cursor.execute(query, args)
        rows = cursor.fetchall()
        numRows = len(list(rows))

        if numRows > 0:
            result = True
    except sqlite3.OperationalError as err:
        print(str(err))
        result = False
        if cursor != None:
            cursor.close()
    finally:
        if cursor != None:
            cursor.close()
        
    return result

def tableExists(table):
    """ Runs a query to see if a table exists """
    result = False
    conn = connectDb()
    try:
        if conn != None:
            qry = "SELECT name from sqlite_master WHERE type='table' AND name=?"
            args = (table,)
            result = coreCursor(conn, qry, args)
            if conn != None:
                conn.close
    except sqlite3.OperationalError as err:
        print(str(err))
        if conn != None:
            conn.close
    return result

def createHashTable():
    """ Creates a DB Table to store hashed files. """
    result = True
    qry = "CREATE TABLE files (file TEXT md5 TEXT)"
    conn = connectDb()
    try:
        if conn != None:
            if not tableExists('files'):
                cursor = conn.cursor()
                try:
                    cursor.execute(qry)
                except sqlite3.OperationalError as err:
                    print(str(err))
                    if cursor != None:
                        cursor.close()
                finally:
                    conn.commit()
                    if cursor != None:
                        cursor.close()
                    result = True
    except sqlite3.OperationalError as err:
        print(str(err))
        if conn != None:
            conn.close()
    finally:
        if conn != None:
            conn.close()
    return result

def createHashtableIdx():
    """ Creates a SQLite DB Table Index. """
    result = False
    qry = "CREATE INDEX idxfile ON files(file, md5)"
    try:
        conn = connectDb()
        if conn != None:
            if tableExists('files'):
                try:
                    cursor = conn.cursor()
                    cursor.execute(qry)
                    result = True
                except sqlite3.OperationalError as err:
                    print(str(err))
                    if cursor != None:
                        cursor.close
    except sqlite3.OperationalError as err:
        print(str(err))
        if conn != None:
            conn.close()
    finally:
        if conn != None:
            conn.close()
    
    return result

def runcmd(qry, args):
    """ Runs specific command on SQLite Database. """
    try:
        conn = connectDb()
        if conn != None:
            if (tableExists()):
                try:
                    cursor = conn.cursor()
                    cursor.execute(qry, args)
                except sqlite3.OperationalError as err:
                    print(str(err))
                    if cursor != None:
                        cursor.close()
    except sqlite3.OperationalError as err:
        print(str(err))
        if conn != None:
            conn.close()
    finally:
        if conn != None:
            conn.close()

def updateHashtable(fname, md5):
    """ Update the SQLite file table. """
    qry = "UPDATE files SET md5=? WHERE file=?"
    args = (md5,fname)
    runcmd(qry, args)

def insertHashtable(fname, md5):
    """ Insert a file and its hash into the Hash Table. """
    args = (fname,md5)
    qry = "INSERT INTO files (file, md5) VALUES(?, ?)"
    runcmd(qry, args)

def setupHashtable(fname, md5):
    """ Set's up the Hash Table. """
    createHashTable()
    createHashtableIdx()
    insertHashtable(fname, md5)

def md5indb(fname):
    """ Checks to see if md5 hash tag exists in database"""
    items = []
    qry = "SELECT md5 WHERE file=? FROM files"
    try:
        conn = connectDb()
        if conn != None:
            if tableExists('files'):
                try:
                    cursor = conn.cursor()
                    cursor.execute(qry,fname)
                    for row in cursor():
                        items.append(row[0])
                except sqlite3.OperationalError as err:
                    print(str(err))
                    if cursor != None:
                        cursor.close
                finally:
                    if cursor != None:
                        cursor.close()
    except sqlite3.OperationalError as err:
        print(str(err))
        if conn != None:
            conn.close()
    finally:
        if conn != None:
            conn.close()
    return items

def hasChanged(fname, md5):
    """ Checks to see if a files md5 has changed"""
    result = False
    oldMd5 = md5indb(fname)
    if len(oldMd5) > 0:
        if md5 != oldMd5:
            result = True
            updateHashtable(fname, md5)
    else:
        setupHashtable(fname, md5)
    return result

def getFileExt(fname):
    """ Returns a file name extension. """
    return os.path.splitext(fname)[1]

def getModDate(fname):
    """ Returns the specified files last modified date. """
    try:
        mtime = os.path.getmtime(fname)
    except OSError as emsg:
        print(str(emsg))
        mtime = 0
    return mtime

def md5Short(fname):
    """ Get md5 file hash tag. """
    return hashlib.md5(str(fname + '|' + str(getModDate(fname))).encode('utf8')).hexdigest()

def getdt(fmt):
    """Get the current DateTime as a string"""
    now = datetime.today()
    dt_string = datetime.strftime(now, fmt)
    return dt_string

def startxlsreport():
    """ Create workbook, get the hostname and datetime """
    wb = workbook.Workbook()
    ws = socket.gethostname()
    st = getdt("%d-%b-%Y %H_%M_%S")
    return wb, ws, st

def endxlsreport(wb, st):
    """ Finalize the excel report. """
    dt = ' from ' + st + ' to ' + getdt("%d-%b-%Y %H_%M_%S")
    fn = basefile() + dt + '.xlsl'
    wb.save(fn)

def headerxlsreport(ws):
    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Full File")
    ws.cell(row=1, column=3, value="Folder Name")
    ws.cell(row=1, column=4, value="Date")
    ws.cell(row=1, column=5, value="Time")

    a1 = ws['A1']
    b1 = ws['B1']
    c1 = ws['C1']
    d1 = ws['D1']
    e1 = ws['E1']

    a1.font = Font(color="00000", bold=True)
    b1.font = Font(color="00000", bold=True)
    c1.font = Font(color="00000", bold=True)
    d1.font = Font(color="00000", bold=True)
    e1.font = Font(color="00000", bold=True)

def getLastRow(ws):
    """ Get the last row in an excel spreadsheet. """
    row = 1
    for cell in ws["A"]:
        if cell.value is None:
            break
        else:
            row += 1
    return row

def rowxlsreport(ws, fn, ffn, fld, d, t):
    rw = getLastRow(ws)

    ws.cell(row=rw, column=1, value=fn)
    ws.cell(row=rw, column=2, value=ffn)
    ws.cell(row=rw, column=3, value=fld)
    ws.cell(row=rw, column=4, value=d)
    ws.cell(row=rw, column=5, value=t)

def loadFlds():
    """ Parse the configuration file and return folders to be included
     and extensions to be excluded. """
    flds = []
    ext = []
    config = basefile() + '.ini'
    if os.path.isfile(config):
        cfile = open(config, 'r')
        # Parse each config file line and get the folder and extensions.
        for line in cfile:
            if '|' in line:
                flds.append(line.split('|')[0])
                exts = line.split('|')[1]
                if len(exts) > 0:
                    ext.append(exts.split(','))
            else:
                flds.append(line)
                ext.append([])
        cfile.close()
    return flds, ext

def checkFileChanges(folder, exclude, ws):
    """ Checks folder structure for files that have changed """
    changed = False
    for subdir, dirs, files, in os.walk(folder):
        for fname in files:
            origin = os.path.join(subdir, fname)
            if os.path.isfile(origin):
                fext = getFileExt(origin)
                if not fext in exclude:
                    md5 = md5Short(origin)
                    headerxlsreport(ws)
                    if hasChanged(origin, md5):
                        now = getdt("%d-%b-%Y %H:%M:%S")
                        dt = now.split(' ')
                        rowxlsreport(ws, fname, origin, subdir, dt[0], dt[1])
                        print(origin + ' has changed on ' + now)
                        changed = True
    return changed

def runFileChanges(ws):
    """ Invoke the function that loads and parses the config file """
    changed = False
    fldExts = loadFlds()
    for i , fld in enumerate(fldExts[0]):
        exts = fldExts[1]
        changed = checkFileChanges(fld, exts[i], ws)       
    return changed

def execute(args):
    chn = 0
    if len(args) > 1:
        if args[1] == '--loop':
            wb, ws, st = startxlsreport()
            try:
                while True:
                    changed = runFileChanges(ws)
                    if changed:
                        chn += 1
            except KeyboardInterrupt:
                print("Stopped...")
                if chn > 0:
                    endxlsreport(wb, st)
    else:
        wb, ws, st = startxlsreport()
        changed = runFileChanges(ws)
        if changed:
            endxlsreport(wb, st)

if __name__ == '__main__':
    execute(sys.argv)
